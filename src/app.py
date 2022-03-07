import sys
import qdarkstyle
import json
import time
from openpyxl import load_workbook, Workbook
from gui import Ui_Form
from PyQt5 import QtCore
from PyQt5.QtGui import QStandardItemModel, QValidator
from PyQt5.QtWidgets import (
    QFileDialog, 
    QWidget, 
    QApplication, 
    QMessageBox, 
    QTableWidgetItem,
    QTableView
)

CONFIG = json.load(open('config.json', 'r'))

WORDS = [w.upper() for w in sorted(list(set(CONFIG['words'].replace(' ','').split(','))))]
CHARS = ''.join(list(CONFIG['chars'].replace(' ','')))


class Validator(QValidator):
    def validate(self, string, pos):
        return QValidator.Acceptable, string.upper(), pos

class MyWindow(QWidget):
 
    def __init__(self):
     
        super(MyWindow, self).__init__()
        self.ui = Ui_Form()
        self.ui.setupUi(self)
        self.ui.button_open.clicked.connect(self.openFileNameDialog)
        self.ui.button_clean.clicked.connect(self.clean_col)
        self.ui.button_save.clicked.connect(self.save_spreadsheet)
        self.validator = Validator()
        self.ui.words_edit.setText(','.join(WORDS))
        self.ui.chars_edit.setText(''.join(CHARS))
        #self.showMaximized()

        self.ui.words_edit.textChanged.connect(self.save_filters)
        self.ui.words_edit.setValidator(self.validator)
        self.ui.chars_edit.textChanged.connect(self.save_filters)

    def get_header_values(self):
        headers = []
        for column in range(self.ui.table_widget.columnCount()):
            header = self.ui.table_widget.horizontalHeaderItem(column)
            headers.append(header.text() or "")
        return headers

    def save_table_to_spreadsheet(self, filename):
        row_count = self.ui.table_widget.rowCount()
        column_count = self.ui.table_widget.columnCount()
        
        wb = Workbook()
        ws = wb.active
        
        #headers
        for col in range(column_count):
            for h in self.get_header_values():
                ws.cell(row=1, column=col+1).value = h        
        
        #rest
        for row in range(row_count):
            data = [self.ui.table_widget.item(row, col).text() for col in range(column_count)]
            for col, val in enumerate(data):
                ws.cell(row=row+2, column=col+1).value = val
        
        wb.save(filename[0]) #this is a tuple with filename and filetype

    def _filter(self, data):
        words = self.ui.words_edit.text()
        chars = self.ui.chars_edit.text()
        
        result = ' '.join([w for w in data.upper().split() if w.upper() not in words])
        result = ''.join([i for i in result if i not in chars])
        return result

    def clean_col(self):
        row_idx = self.ui.table_widget.rowCount()
        words = self.ui.words_edit.text().split(',')
        chars = list(self.ui.chars_edit.text())
        indexes = self.ui.table_widget.selectionModel().selectedColumns()
        cols = [index.column() for index in sorted(indexes)]
        for col in cols:
            for row in range(row_idx):
                data = self.ui.table_widget.item(row, col).text()
                self.ui.table_widget.item(row, col).setText(self._filter(data))
                
        self.ui.table_widget.resizeColumnsToContents()
        
        msg = QMessageBox() 
        msg.setText("Palavras e caracteres removidos com sucesso.")
        msg.exec_()

    def save_filters(self):
        words = self.ui.words_edit.text()
        chars = self.ui.chars_edit.text()
        open('config.json','w').write(json.dumps(dict(words=words, chars=chars)))

    def insert_row(self, *args):
        row_idx = self.ui.table_widget.rowCount()
        tbl = self.ui.table_widget
        tbl.insertRow(row_idx)
        for col_idx, arg in enumerate(args):
            tbl.setItem(row_idx, col_idx, QTableWidgetItem(arg))

    def load_spreadsheet(self, filename):
        wb = load_workbook(filename)
        sheet = wb.worksheets[0]
        row_count = sheet.max_row
        column_count = sheet.max_column
        self.ui.table_widget.setColumnCount(column_count)
        self.ui.table_widget.setRowCount(0)
        
        headers = [sheet.cell(row=1, column=col+1).value for col in range(column_count)]
        self.ui.table_widget.setHorizontalHeaderLabels(headers)
        self.ui.table_widget.horizontalHeader().setDefaultAlignment(QtCore.Qt.AlignLeft)
        
        for row in range(1, row_count):
            contents = [str(sheet.cell(row=row+1, column=col+1).value) for col in range(column_count)]
            self.insert_row(*contents)
            
        self.ui.table_widget.resizeColumnsToContents()
            

    def openFileNameDialog(self):
        filename, _ = QFileDialog.getOpenFileName(self, "", "", "Planilha Excel (*.xlsx)")
        if filename:
            try:
                self.load_spreadsheet(filename)
            except Exception as e:
                msg = QMessageBox() 
                msg.setText(str(e))
                msg.setIcon(QMessageBox.Critical)
                msg.exec_()
                
    def save_spreadsheet(self):
        filename = QFileDialog.getSaveFileName(self, "", "", "Planilha Excel (*.xlsx)")
        try:
            self.save_table_to_spreadsheet(filename)
            msg = QMessageBox() 
            msg.setText("Arquivo salvo com sucesso!")
            msg.exec_()
        except Exception as e:
            msg = QMessageBox() 
            msg.setText(str(e))
            msg.setIcon(QMessageBox.Critical)
            msg.exec_()

app = QApplication([])
application = MyWindow()
application.show()
app.setStyleSheet(qdarkstyle.load_stylesheet_pyqt5())
sys.exit(app.exec())